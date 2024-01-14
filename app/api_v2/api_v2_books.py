from datetime import date, timedelta
from sqlalchemy import update
import openpyxl
from flask import jsonify, request ,send_file,after_this_request
import random
import json
from random import shuffle
from sqlalchemy import Date, Integer, and_, cast, desc, or_, table,desc, asc, nullslast
from sqlalchemy.inspection import inspect
from app import db
from app.models.new_books import NewBookImage,NewBookVideo, NewBookSection, NewBook, NewCategory, NewCategoryBook
from app.models.order import Order
from app.models.buckets import Wishlist, Dump
from app.models.books import Book
import datetime
import xlsxwriter
from app.api_admin.utils import upload_to_aws
from app.api_v2.utils import api_v2_books
from sqlalchemy import func
import os
import csv



@api_v2_books.route('/get-book-set')
def get_book_set(): 
    age = request.args.get('age')
    section_name = request.args.get('section_name')
    start = request.args.get('start')
    end = request.args.get('end')
    if not section_name or age is None: 
        return jsonify({"success": False, "message": "Provide age group and section name"})
    if not str(age).isnumeric(): 
        return jsonify({"success": False, "message": "Invalid age group"})
    if start and str(start).isnumeric(): 
        start = int(start)
    else: 
        start = None
    if end and str(end).isnumeric(): 
        end = int(end)
    else: 
        end = None
    section = NewBookSection.query.filter_by(name=section_name).first()
    if not section: 
        return jsonify({"success": False, "message": "Invalid section name"})
    age = int(age)
    categories_query = NewCategory.query.filter(
        NewCategory.min_age <= age,
        NewCategory.max_age >= age
    ).order_by(func.random())
    
    if start==0:
     best_seller_category = NewCategory.query.filter_by(name='Best Seller - Most Popular').first()
    if start is not None and end is not None: 
        categories_query = categories_query.limit(end - start).offset(start)
        
    categories = categories_query.all()
    if start==0:
     categories.append(best_seller_category)
   
    book_set = []
    # if len(categories) > 1: 
    #     shuffled_categories = categories[1:]
    #     random.shuffle(shuffled_categories)
    #     categories = [categories[0], *shuffled_categories]
    for category in categories: 
        books = db.session.query(NewBook, NewCategoryBook).filter(
            NewBook.id == NewCategoryBook.book_id,
            NewCategoryBook.category_id == category.id,
            NewCategoryBook.section_id == section.id,
            NewBook.min_age <= age, 
            NewBook.max_age >= age
        ).all()
        books = [book[0].to_json() for book in books]
        # for book in books:
        #  if isinstance(book['review_count'], str):
        #     book['review_count'] = book['review_count'].replace(',', '')
        # if category.name == 'Best Seller - Most Popular': 
        #     random.shuffle(books)
        # else: 
        #     books = sorted(books, key=lambda book: int(book['review_count']) if isinstance(book['review_count'], int) or book['review_count'].isdigit() else 0, reverse=True)
        if len(books): 
            book_set.append({
                "category": category.name,
                "books": books
            })
        # random.shuffle(book_set)    
    return jsonify({"success": True, "book_set": book_set})

@api_v2_books.route('/get-most-popular-set')
def get_most_popular_set(): 
    age = request.args.get('age')
    count = request.args.get('count')
    category_limit = request.args.get('category_limit')
    if not str(age).isnumeric(): 
        return jsonify({"success": False, "message": "Invalid age group"})
    if not count or not count.isnumeric() or int(count) <= 0: 
        return jsonify({"success": False, "message": "Invalid count"})
    if not category_limit or not category_limit.isnumeric() or int(category_limit) <= 0: 
        return jsonify({"success": False, "message": "Invalid category limit"})
    age = int(age)
    count = int(count)
    category_limit = int(category_limit)
    categories = NewCategory.query.filter(
        NewCategory.min_age <= age, 
        NewCategory.max_age >= age
    ).all()
    books = []
    book_ids = set()
    for category in categories: 
        category_books = db.session.query(NewBook).join(NewCategoryBook, NewCategory).filter(
            NewBook.id == NewCategoryBook.book_id,
            NewCategoryBook.category_id == category.id,
            NewBook.min_age <= age, 
            NewBook.max_age >= age,
        ).join(Book, NewBook.isbn == Book.isbn).filter(
            Book.stock_available > 0,
        ).order_by(desc(cast(NewBook.review_count, Integer))).limit(category_limit).all()
        for book in category_books: 
            if book.id not in book_ids: 
                books.append(book)
                book_ids.add(book.id)
    books = sorted(books, key=lambda book: int(book.review_count), reverse=True)[:count]
    books = [book.to_json() for book in books]
    return jsonify({"success": True, "books": books})

@api_v2_books.route('/get-must-read-set')
def get_must_read_set(): 
    age = request.args.get('age')
    category_count = request.args.get('category_count')
    book_count = request.args.get('book_count')
    section_name = request.args.get('section_name')
    show_unavailable = request.args.get('show_unavailable')
    randomize_categories = request.args.get('randomize_categories')
    randomize_books = request.args.get('randomize_books')
    if not str(age).isnumeric(): 
        return jsonify({"success": False, "message": "Invalid age group"}), 400
    if not category_count or not book_count: 
        return jsonify({"success": False, "message": "Provide category and book count"}), 400
    if not category_count.isnumeric() or int(category_count) <= 0: 
        return jsonify({"success": False, "message": "Invalid category count"}), 400
    if not book_count.isnumeric() or int(book_count) <= 0: 
        return jsonify({"success": False, "message": "Invalid book count"}), 400
    section = NewBookSection.query.filter_by(name=section_name).first()
    if not section: 
        return jsonify({"success": False, "message": "Invalid section name"}), 400
    age = int(age)
    category_count = int(category_count)
    book_count = int(book_count)
    categories = NewCategory.query.join(NewCategoryBook).filter(NewCategoryBook.section_id == section.id).all()
    book_set = []
    if len(categories) == 1 and categories[0].name == section.name: 
        books_query = NewBook.query.join(NewCategoryBook).filter(
            NewBook.id == NewCategoryBook.book_id,
            NewCategoryBook.section_id == section.id,
            NewCategoryBook.category_id == categories[0].id,
            NewBook.min_age <= age, 
            NewBook.max_age >= age
        )
        if not show_unavailable: 
            books_query = books_query.join(Book, NewBook.isbn == Book.isbn).filter(Book.stock_available >= 1)
        books = books_query.order_by(desc(cast(NewBook.review_count, Integer))).all()
        category_to_books = dict()
        for book in books: 
            categories = NewCategory.query.join(NewCategoryBook).filter(
                NewCategoryBook.section_id == 1,
                NewCategoryBook.book_id == book.id,
                NewCategory.max_age != 100,
            ).all()
            for category in categories: 
                if category.name not in category_to_books: 
                    category_to_books[category.name] = {"category": category.name, "books": []}
                if len(category_to_books[category.name]["books"]) < book_count: 
                    category_to_books[category.name]["books"].append(book)
        for category in category_to_books: 
            if randomize_books: 
                random.shuffle(category_to_books[category]["books"])
            book_set.append(category_to_books[category])
    else: 
        for category in categories: 
            books_query = NewBook.query.join(NewCategoryBook, NewCategory).filter(
                NewBook.id == NewCategoryBook.book_id,
                NewCategoryBook.category_id == category.id,
                NewBook.min_age <= age, 
                NewBook.max_age >= age
            )
            if not show_unavailable: 
                books_query = books_query.join(Book, NewBook.isbn == Book.isbn).filter(Book.stock_available >= 1)
            books_all = books_query.order_by(desc(cast(NewBook.review_count, Integer))).all()
            books = []
            category_ids = set()
            for book in books_all: 
                book_categories = NewCategoryBook.query.filter_by(book_id=book.id, section_id=1).all()
                category_exists = False
                for book_category in book_categories: 
                    if book_category.category_id in category_ids: 
                        category_exists = True
                        break
                    category_ids.add(book_category.category_id)
                if not category_exists: 
                    books.append(book)
                if len(books) == book_count: 
                    break
            if randomize_books: 
                random.shuffle(books)
            book_set.append({"category": category.name, "books": books})
    book_set = sorted(book_set, key=lambda category: sum([int(book.review_count) for book in category["books"]]), reverse=True)[:category_count]
    if randomize_categories: 
        random.shuffle(book_set)
    for i in range(category_count): 
        book_set[i]["books"] = [book.to_json() for book in book_set[i]["books"]]
    return jsonify({"success": True, "book_set": book_set})

@api_v2_books.route('/get-category-books')
def get_category_books(): 
    category_name = request.args.get('category_name')
    if not category_name: 
        return jsonify({"success": False, "message": "Provide category name"}), 400
    category = NewCategory.query.filter_by(name=category_name).first()
    if not category: 
        return jsonify({"success": False, "message": "Invalid category name"}), 400
    books = category.books
    random.shuffle(books)
    return jsonify({"success": True, "books": [book.to_json() for book in books]})

@api_v2_books.route('/get-new-books')
def get_new_books(): 
    start = request.args.get('start') 
    end = request.args.get('end')
    age = request.args.get('age')
    category_id = request.args.get('category_id')
    search_query = request.args.get('search_query')
    section_id = request.args.get('section_id')
    sort_review_count = request.args.get('sort_review_count')
    
    if age and not age.isnumeric() and age != '-1': 
        return jsonify({"success": False, "message": "Provide a valid age group"}), 400
    
    if category_id and not NewCategory.query.filter_by(id=category_id).first(): 
        return jsonify({"success": False, "message": "Invalid category ID"}), 400
    
    if section_id and not NewBookSection.query.filter_by(id=section_id).first(): 
        return jsonify({"success": False, "message": "Invalid section ID"}), 400
    
    if not search_query: 
        search_query = ''
    
    if not start or not start.isnumeric(): 
        start = 0
    
    if not end or not end.isnumeric(): 
        end = 10
    
    age = int(age) if age and age != '-1' else None
    start, end = int(start), int(end)
    
    books_query = db.session.query(NewBook)
    
    if search_query:
        # Use outerjoin with the appropriate onclause
        books_query = books_query.outerjoin(
            NewCategoryBook, 
            NewCategoryBook.book_id == NewBook.id  # Correct the onclause
        ).outerjoin(
            NewCategory,
            NewCategory.id == NewCategoryBook.category_id  # Correct the onclause
        ).filter(
            or_(
                NewBook.name.ilike(f'{search_query}%'),
                NewBook.isbn.ilike(f'{search_query}%'),
                NewCategory.name.ilike(f'{search_query}%'),
            )
        )
    
    if age is not None: 
        books_query = books_query.filter(
            NewBook.min_age <= age, 
            NewBook.max_age >= age
        )
    
    if category_id and section_id: 
        books_query = books_query.filter(
            NewBook.id == NewCategoryBook.book_id,
            NewCategoryBook.category_id == category_id,
            NewCategoryBook.section_id == section_id,
        )
    elif category_id: 
        books_query = books_query.filter(
            NewBook.id == NewCategoryBook.book_id,
            NewCategoryBook.category_id == category_id
        )
    elif section_id: 
        books_query = books_query.filter(
            NewBook.id == NewCategoryBook.book_id,
            NewCategoryBook.section_id == section_id
        )
    
    # Sorting based on your criteria
    books_query = books_query.order_by(
        nullslast(desc(NewBook.book_order)),
        nullslast(asc(NewBook.publication_date)),
        asc(NewBook.review_count)
    )
    
    books = []
    
    for book in books_query.limit(end - start).offset(start).all():
        return_date = None
        
        old_book = Book.query.filter_by(isbn=book.isbn).first()
        
        if old_book is not None:
           if not old_book.stock_available: 
             order = Order.query.filter(
                Order.book_id == old_book.id,
                cast(Order.placed_on, Date) >= cast(date.today() + timedelta(days=-7), Date)
             ).order_by(Order.placed_on).first()
            
             if order:
                return_date = order.placed_on + timedelta(days=7)
        
           wishlist_count = Wishlist.query.filter_by(book_id=old_book.id).count()
           previous_count = Dump.query.filter_by(book_id=old_book.id, read_before=True).count() + \
                         Order.query.filter_by(book_id=old_book.id).count()
            
           books.append({
              **book.to_json(),   
              "return_date": return_date,
              "wishlist_count": wishlist_count,
              "previous_count": previous_count
            })
    
    book_ids_in_array = [book['id'] for book in books]   
    for book in books_query:
  
     if book.id not in book_ids_in_array:
        rentals = book.rentals
        stock_available = book.stock_available 
        
        books.append({
            **book.to_json(),
            "rentals": rentals,
            "stock_available": stock_available
        })
            
    return jsonify({"success": True, "books": books})

@api_v2_books.route('/search-new-books')
def search_new_books(): 
    search_query = request.args.get('search_query')
    start = request.args.get('start')
    end = request.args.get('end')
    if not search_query or not search_query.strip(): 
        return jsonify({"success": False, "message": "Enter a search query"}), 400
    if len(search_query) < 3: 
        return jsonify({"success": False, "message": "Search query should be atleast 3 characters long"}), 400
    if not start or not start.isnumeric(): 
        start = 0
    if not end or not end.isnumeric(): 
        end = 100
    start, end = int(start), int(end)
    books = NewBook.query.join(NewCategoryBook, NewCategory).filter(
        NewCategoryBook.category_id == NewCategory.id,
        NewCategory.name != 'Best Seller - Most Popular',
        NewCategoryBook.section_id == 1,
        or_(
            NewBook.name.ilike(f'% {search_query} %'),
            NewBook.isbn.ilike(f'%{search_query}%'),
            NewBook.authors.ilike(f'%{search_query}%'),
            NewBook.book_type.ilike(f'%{search_query}%'),
            NewBook.publisher.ilike(f'%{search_query}%'),
            NewBook.description.ilike(f'% {search_query} %'),
            NewCategory.name.ilike(f'%{search_query}%'),
        )
    ).limit(end - start).offset(start).all()
    category_to_books = dict()
    for book in books: 
        categories = NewCategory.query.join(NewCategoryBook).filter(
            NewCategoryBook.section_id == 1,
            NewCategoryBook.book_id == book.id,
            NewCategory.max_age != 100,
        ).all()
        for category in categories: 
            if category.name not in category_to_books: 
                category_to_books[category.name] = {
                    "category": category.name, 
                    "books": [],
                    "min_age": category.min_age,
                    "max_age": category.max_age,
                }
            category_to_books[category.name]["books"].append(book.to_json())
    book_set = [category_to_books[category] for category in category_to_books]
    book_set = sorted(book_set, key=lambda books: len(books["books"]), reverse=True)
    return jsonify({"success": True, "book_set": book_set})

@api_v2_books.route('/new-book', methods=['POST', 'PUT'])
def new_book(): 
    id = request.form.get('id')
    isbn = request.form.get('isbn')
    name = request.form.get('name')
    min_age = request.form.get('min_age')
    max_age = request.form.get('max_age')
    image = request.form.get('image')
    rating = request.form.get('rating')
    review_count = request.form.get('review_count')
    categories = request.form.get('categories')
    genre = request.form.get('genre')
    pages = request.form.get('pages')
    lexile_measure = request.form.get('lexile_measure')
    description = request.form.get('description')
    publication_date = request.form.get('publication_date')
    publisher = request.form.get('publisher')
    author = request.form.get('author')
    language = request.form.get('language')
    paperbackprice = request.form.get('paperbackprice')
    boardbookprice = request.form.get('boardbookprice')
    hardcoverprice = request.form.get('hardcoverprice')
    
    
    image_file = request.files.get('image')
    if not all((isbn, name, min_age, max_age, rating, review_count, categories)) or (not image and not image_file): 
        return jsonify({"success": False, "message": "Provide all the data"}), 400
    if not str(min_age).isnumeric() or not str(max_age).isnumeric() or int(min_age) < 0 or int(min_age) > int(max_age): 
        return jsonify({"success": False, "message": "Invalid minimum and maximum age"}), 400
    if not str(rating).replace('.', '').isnumeric() or float(rating) < 0 or float(rating) > 5: 
        return jsonify({"success": False, "message": "Invalid rating"}), 400
    if not str(review_count).isnumeric() or int(review_count) < 0: 
        return jsonify({"success": False, "message": "Invalid review count"}), 400
    if type(json.loads(categories)) != type([]): 
        return jsonify({"success": False, "message": "Invalid category list"}), 400
    
    if pages == 'null':
            pages=None
    if description == 'null':
            description=None
    if lexile_measure == 'null':
            lexile_measure=None
    if publisher == 'null':
            publisher=None
    if genre == 'null':
            genre=None
    if author == 'null':
            author=None 
    if language == 'null':
            language=None 
    if publication_date == 'null':
            publication_date=None 
    if paperbackprice == 'null':
            paperbackprice=None 
    if hardcoverprice == 'null':
            hardcoverprice=None  
    if boardbookprice == 'null':
            boardbookprice=None 
   

    min_age = int(min_age)
    max_age = int(max_age)
    categories = json.loads(categories)
    if request.method == 'POST': 
        if NewBook.query.filter_by(isbn=isbn).count(): 
            return jsonify({"success": False, "message": "Book with given ISBN already exists"}), 400
        if NewBook.query.filter_by(isbn=isbn).count(): 
            return jsonify({"success": False, "message": "Book with given ISBN already exists"}), 400
        
        book_image = image
        if not image or not image.startswith('http'): 
            extension = image_file.filename.split(".")[-1]
            upload_to_aws(image_file, 'book_images', f'book_images/{isbn}.{extension}')
            s3_url = os.environ.get('AWS_S3_URL')
            book_image = f'{s3_url}/book_images/{isbn}.{extension}'
            
       
                
            
        NewBook.create(
            name, 
            book_image, 
            isbn, 
            rating, 
            review_count, 
            min_age, 
            max_age,
            language,
            genre,
            pages,
            lexile_measure,
            description,
            publisher,
            publication_date,
            paperbackprice,
            boardbookprice,
            hardcoverprice,
            author
        )
        new_book = NewBook.query.filter_by(isbn=isbn).first()

        for category in categories: 
            NewCategoryBook.create(
                category['category']['id'],
                new_book.id,
                category['section']['id'],
            )

        if not Book.query.filter_by(isbn=isbn).count():
            Book.create(
                name, 
                book_image, 
                isbn, 
                rating, 
                review_count, 
                '', 
                'English', 
                '',
                '', 
                1, 
                None, 
                None, 
                None,
                None
            )
        book = Book.query.filter_by(isbn=isbn).first()
        book.age_group_1 = (min_age >= 0 and min_age <= 2) or (max_age >= 0 and max_age <= 2)
        book.age_group_2 = (min_age >= 3 and min_age <= 5) or (max_age >= 3 and max_age <= 5)
        book.age_group_3 = (min_age >= 6 and min_age <= 8) or (max_age >= 6 and max_age <= 8)
        book.age_group_4 = (min_age >= 9 and min_age <= 11) or (max_age >= 9 and max_age <= 11)
        book.age_group_5 = (min_age >= 12 and min_age <= 14) or (max_age >= 12 and max_age <= 14)
        book.age_group_6 = (min_age >= 15) or (max_age >= 15)

        db.session.commit()
        
        return jsonify({"success": True, "book": new_book.to_json()})
    elif request.method == 'PUT': 
        new_book = NewBook.query.filter_by(id=id).first()
        
        if not new_book: 
            return jsonify({"success": False, "message": "Invalid book ID"}), 404
        
        book = Book.query.filter_by(isbn=new_book.isbn).first()

        book_image = image
        if not image or not image.startswith('http'): 
            extension = image_file.filename.split(".")[-1]
            upload_to_aws(image_file, 'book_images', f'book_images/{isbn}.{extension}')
            s3_url = os.environ.get('AWS_S3_URL')
            book_image = f'{s3_url}/book_images/{isbn}.{extension}'
        
        
        if not new_book:
          book.isbn = isbn
          book.name = name
          book.image = book_image
          book.review_count = review_count
          book.rating = rating
          new_book.min_age = min_age
          new_book.max_age = max_age
          new_book.genre=genre
          if lexile_measure is not None:
           new_book.lexile_measure = lexile_measure

          if pages is not None and pages != '':
           print (pages)  
           pages=int(pages)
           new_book.pages = pages
           
          if paperbackprice is not None and paperbackprice != '':
           print (paperbackprice)  
           paperbackprice=int(paperbackprice)
           new_book.paperbackprice = paperbackprice 
         
          if boardbookprice is not None and boardbookprice != '':
           print (boardbookprice)  
           boardbookprice=int(boardbookprice)
           new_book.boardbookprice = boardbookprice   
          
          if hardcoverprice is not None and hardcoverprice != '':
           print (hardcoverprice)  
           hardcoverprice=int(hardcoverprice)
           new_book.hardcoverprice = hardcoverprice    

           if publisher is not None:
            new_book.publisher = publisher

           if publication_date is not None and publication_date != 'null':
               
            new_book.publication_date = publication_date

           if language is not None:
            new_book.language = language
            
           if author is not None:
            new_book.author = author 

           if description is not None:
            book.description = description
        else:  
          new_book.isbn =  isbn
          new_book.name =  name
          new_book.image = book_image
          new_book.review_count = review_count
          new_book.rating  = rating
          new_book.min_age = min_age
          new_book.max_age = max_age
          new_book.genre=genre
          if lexile_measure is not None:
           new_book.lexile_measure = lexile_measure

          if pages is not None and pages != '':
           print(pages)   
           pages=int(pages)
           new_book.pages = pages
          
          if paperbackprice is not None and paperbackprice != '':
           print (paperbackprice)  
           paperbackprice=int(paperbackprice)
           new_book.paperbackprice = paperbackprice 
         
          if boardbookprice is not None and boardbookprice != '':
           print (boardbookprice)  
           boardbookprice=int(boardbookprice)
           new_book.boardbookprice = boardbookprice   
          
          if hardcoverprice is not None and hardcoverprice != '':
           print (hardcoverprice)  
           hardcoverprice=int(hardcoverprice)
           new_book.hardcoverprice = hardcoverprice 

           if publisher is not None:
            new_book.publisher = publisher

           if publication_date is not None and publication_date != 'null':
            new_book.publication_date = publication_date

           if language is not None:
            new_book.language = language
            
           if author is not None:
            new_book.author = author   

           if description is not None:
            new_book.description = description
          
          

        for category in NewCategoryBook.query.filter_by(book_id=new_book.id).all(): 
            category.delete()
        for category in categories: 
            print(category)
            NewCategoryBook.create(
                category['category']['id'],
                new_book.id,
                category['section']['id'],
            )
            print(NewCategoryBook)
            
        
        if book:
         book.age_group_1 = (min_age >= 0 and min_age <= 2) or (max_age >= 0 and max_age <= 2)
         book.age_group_2 = (min_age >= 3 and min_age <= 5) or (max_age >= 3 and max_age <= 5)
         book.age_group_3 = (min_age >= 6 and min_age <= 8) or (max_age >= 6 and max_age <= 8)
         book.age_group_4 = (min_age >= 9 and min_age <= 11) or (max_age >= 9 and max_age <= 11)
         book.age_group_5 = (min_age >= 12 and min_age <= 14) or (max_age >= 12 and max_age <= 14)
         book.age_group_6 = (min_age >= 15) or (max_age >= 15)

        db.session.commit()

        return jsonify({"success": True, "book": new_book.to_json()})

@api_v2_books.route('/update-book-quantity', methods=['POST'])
def update_book_quantity(): 
    id = request.json.get('id')
    stock_available = request.json.get('stock_available')
    rentals = request.json.get('rentals')
    if not str(stock_available).isnumeric() or not str(rentals).isnumeric() or int(stock_available) < 0 or int(rentals) < 0: 
        
        return jsonify({"success": False, "message": "Invalid book quantity"}), 400
    new_book = NewBook.query.filter_by(id=id).first()
   
    if new_book is not None: 
        print("newbook")
       
    if not new_book: 
        
        return jsonify({"success": False, "message": "Invalid book ID"}), 404
    book = Book.query.filter_by(isbn=new_book.isbn).first()
    
    if book is not None: 
         print("oldbook")
         
    if not book: 
        
        new_book.stock_available = stock_available
        new_book.rentals = rentals
        
        db.session.commit()
       
        book_json = new_book.to_json()
        book_json["rentals"] = rentals
        book_json["stock_available"] = stock_available
        return jsonify({"success": True, "book": book_json})
   
    book.stock_available = stock_available
    book.rentals = rentals
    
    db.session.commit()
    
    return jsonify({"success": True, "book": new_book.to_json()})

@api_v2_books.route('/delete-new-book', methods=['POST'])
def delete_new_book(): 
    id = request.json.get('id')
    new_book = NewBook.query.filter_by(id=id).first()
    if not new_book: 
        return jsonify({"success": False, "message": "Invalid book ID"}), 404
    new_book.delete()
    return jsonify({"success": True})

@api_v2_books.route('/add-books-from-csv', methods=['POST'])
def add_books_from_csv():
    books_csv_file = request.files.get('books_csv')
    accepted_books = {'csv', 'xlsx'}
    ext = books_csv_file.filename.split(".")[-1]
    if ext not in accepted_books:
        return jsonify({"status": "success", "message": "Only CSV files are supported"}), 400
    if ext == 'csv':
        filename = './temporary.csv'
    else:
        filename = './temporary.xlsx'
    books_csv_file.save(filename)
    books = []
    if ext == 'csv':
        with open(filename, mode="r", encoding='utf-8-sig') as file:
            iterator = csv.reader(file)
            header_row = next(iterator)
            header_row = {i: header_row[i].lower() for i in range(len(header_row))}
            for line in iterator:
                temp = {}
                for value in range(len(list(line))):
                    temp[header_row[value]] = line[value]
                if temp:
                    books.append(temp)
    else:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        iterator = sheet.iter_rows()
        header_row = next(iterator)
        header_row = {i: header_row[i].value.lower() for i in range(len(header_row))}
        for row in iterator:
            temp = {}
            for cell in range(len(row)):
                if row[cell].value:
                    temp[header_row[cell]] = row[cell].value
            if temp:
                books.append(temp)
    if len(books) < 2:
        return jsonify({"status": "success", "message": "Empty CSV file"}), 400
    if 'isbn' not in header_row.values() or 'name' not in header_row.values():
        return jsonify({"status": "success", "message": "CSV file should have atleast ISBN and name column"}), 400
    added_isbns, not_added_isbns = [], []

    inspect_new_book_object = inspect(NewBook)
    inspect_book_object = inspect(Book)
    new_book_columns = [c_attr.key for c_attr in inspect_new_book_object.mapper.column_attrs]
    book_columns = [c_attr.key for c_attr in inspect_book_object.mapper.column_attrs]
    for book in books:
        for x in book:
            # Convert Float and Numeric String --> Integer
            if isinstance(book[x], float) or (isinstance(book[x], str) and book[x].isnumeric()):
                book[x] = int(book[x])

        if 'isbn' not in book or 'name' not in book:
            continue
        book['isbn'] = str(book['isbn'])
        isbn = str(book['isbn'])

        if book['price'] == 'null':
            book['price'] = 0.0
        elif isinstance(book['price'], str):
            book['price'] = float(book['price'].replace(",", ""))
            
        book_attr = {key: value for key, value in book.items() if key in book_columns}
        new_book_attr = {key: value for key, value in book.items() if key in new_book_columns}
        fetch_book = NewBook.query.filter_by(isbn=isbn).first()
        if fetch_book:
            update_new_book = update(NewBook).where(NewBook.isbn == isbn).values(**new_book_attr).returning(NewBook.id)
            updated_id = db.engine.execute(update_new_book).fetchone()
            fetch_category_book = NewCategoryBook.query.filter_by(book_id=updated_id[0]).first()
            db.engine.execute(update(NewCategory).where(NewCategory.id == fetch_category_book.category_id).values(name=book['categories'], category_order=book['order'], min_age=book['min_age'], max_age=book['max_age']))
            update_book = update(Book).where(Book.isbn == isbn).values(**book_attr)
            db.engine.execute(update_book)
            added_isbns.append(isbn)
            print("UPDATED", isbn)
        else:
            new_book_instance = NewBook.create(**new_book_attr)
            book_instance = Book.create(**book_attr)
            category_id = NewCategory.create(book['categories'], book['order'] or 0, book['min_age'], book['max_age'])
            NewCategoryBook.create(category_id, new_book_instance, 1)
            added_isbns.append(isbn)
            print("ADDED", isbn)
        db.session.commit()
    try:
        os.remove(filename)
    except Exception as e:
        print(e)

    return jsonify({"status": "success", "added_isbns": added_isbns, "not_added_isbns": not_added_isbns})

@api_v2_books.route('/getBookAuthor', methods=['GET'])
def get_book_author():
    isbn = request.args.get('isbn')

    if not isbn:
        return jsonify({'error': 'ISBN not provided'}), 400

    # Query the database to find the book with the provided ISBN.
    book = NewBook.query.filter_by(isbn=isbn).first()

    if not book:
        return jsonify({'error': 'Book not found'}), 404

    # Split the authors by comma and create a list of author names.
    authors = book.authors.split(', ')

    # Query the database to find books with at least one of the authors in the list.
    related_books = NewBook.query.filter(NewBook.authors.in_(authors)).all()

    # Create a list of book details (author, rating, review count, ISBN, description).
    book_details = [
        {
            'name':book.name,
            'author': book.authors,
            'rating': book.rating,
            'review_count': book.review_count,
            'isbn': book.isbn,
            'description': book.description,
            'image':book.image,
            "paperbackprice":book.paperbackprice,
            "boardbookprice":book.boardbookprice,
            "hardcoverprice":book.hardcoverprice,
        }
        for book in related_books
    ]

    return jsonify({'author': authors, 'related_books': book_details})

@api_v2_books.route('/getBooksByCategory', methods=['GET'])
def get_books_by_category():
    category_name = request.args.get('category_name')

    if not category_name:
        return jsonify({'error': 'Category name not provided'}), 400

    # Query the database to find the category by name.
    category = NewCategory.query.filter_by(name=category_name).first()

    if not category:
        return jsonify({'error': 'Category not found'}), 404

    # Query the database to find books with the specified category.
    books_in_category = NewBook.query.join(
        NewCategoryBook, NewCategoryBook.book_id == NewBook.id
    ).filter(NewCategoryBook.category_id == category.id).all()

    # Create a list of book details (name, author, rating, review count, ISBN, description).
    book_details = []
    for book in books_in_category:
        authors = book.authors.split(', ') if book.authors else []  # Handle 'authors' attribute gracefully
        book_details.append({
            'name': book.name,
            'authors': authors,
            'rating': book.rating,
            'review_count': book.review_count,
            'isbn': book.isbn,
            'description': book.description,
            'image': book.image,
            'book_order': book.book_order,
            "paperbackprice":book.paperbackprice,
            "boardbookprice":book.boardbookprice,
            "hardcoverprice":book.hardcoverprice,
            'publication_date': book.publication_date.strftime('%Y-%m-%d') if book.publication_date else None
        })

    book_details = sorted(book_details, key=lambda x: (x['book_order'], x['publication_date'] or ''))

    return jsonify({'category_name': category_name, 'books_in_category': book_details})

@api_v2_books.route('/getBookDetails', methods=['GET'])
def get_book_details():
    isbn = request.args.get('isbn')

    if not isbn:
        return jsonify({'error': 'ISBN not provided'}), 400

    # Query the database to find the book with the provided ISBN.
    book = NewBook.query.filter_by(isbn=isbn).first()

    if not book:
        return jsonify({'error': 'Book not found'}), 404

    # Create a dictionary with all book details.
    book_details = {
        'name': book.name,
        'author': book.authors,
        'rating': book.rating,
        'review_count': book.review_count,
        'isbn': book.isbn,
        'description': book.description,
        'for_age': book.for_age,
        'grade_level': book.grade_level,
        'lexile_measure': book.lexile_measure,
        'pages': book.pages,
        'dimensions': book.dimensions,
        'publisher': book.publisher,
        'publication_date': book.publication_date.strftime('%Y-%m-%d') if book.publication_date else None,
        'language': book.language,
        "paperbackprice":book.paperbackprice,
        "boardbookprice":book.boardbookprice,
        "hardcoverprice":book.hardcoverprice,
        'genre':book.genre,
    }

    return jsonify({'book_details': book_details})

@api_v2_books.route('/getAuthorsByISBN', methods=['GET'])
def get_authors_by_isbn():
    isbn = request.args.get('isbn')

    if not isbn:
        return jsonify({'error': 'ISBN not provided'}), 400

    # Query the database to find the book with the provided ISBN.
    book = NewBook.query.filter_by(isbn=isbn).first()

    if not book:
        return jsonify({'error': 'Book not found'}), 404

    # Split the authors by comma and create a list of author names.
    authors = book.authors.split(', ')

    return jsonify({'isbn': isbn, 'authors': authors})

@api_v2_books.route('/getBooksByAuthor', methods=['GET'])
def get_books_by_author():
    author_name = request.args.get('author')

    if not author_name:
        return jsonify({'error': 'Author name not provided'}), 400

    # Query the database to find books with the provided author's name.
    author_name = author_name.strip()  # Remove leading/trailing spaces
    books = NewBook.query.filter(NewBook.authors.ilike(f'%{author_name}%')).all()

    if not books:
        return jsonify({'error': 'No books found for the author'}), 404

    # Create a list of book details with author names as an array.
    book_details = [
        {
            'name': book.name,
            'authors': book.authors.split(', '),  # Convert authors to an array
            'isbn': book.isbn,
            'rating': book.rating,
            'review_count': book.review_count,
            'description': book.description,
            'image': book.image,
            "paperbackprice":book.paperbackprice,
            "boardbookprice":book.boardbookprice,
            "hardcoverprice":book.hardcoverprice,
        }
        for book in books
    ]
    book_details = sorted(book_details, key=lambda x: int(x['review_count']), reverse=True)

    return jsonify({'author': author_name, 'books': book_details})

@api_v2_books.route('/getTopBooksByReviewCount', methods=['GET'])
def get_top_books_by_review_count():
    age = request.args.get('age')

    if not age or not age.isnumeric():
        return jsonify({'error': 'Invalid or missing age parameter'}), 400

    age = int(age)

    # Query the database to find books within the specified age group and order them by review_count.
    books = NewBook.query.filter(NewBook.min_age <= age, NewBook.max_age >= age) \
        .order_by(NewBook.review_count.desc()).limit(30).all()

    # Create a list of book details (name, author, rating, review count, ISBN, description).
    book_details = []
    
    for book in books:
        authors = book.authors.split(', ') if book.authors else []  # Handle 'authors' attribute gracefully
        stock_available = 0
        book_record = Book.query.filter_by(isbn=book.isbn).first()
        if book_record:
            stock_available = book_record.stock_available
        else:
            stock_available=book.stock_available
        
        if stock_available:
            
         book_details.append({
            'name': book.name,
            'authors': authors,
            'rating': book.rating,
            'review_count': book.review_count,
            'isbn': book.isbn,
            'description': book.description,
            'image': book.image,
            'book_order': book.book_order,
            'publication_date': book.publication_date.strftime('%Y-%m-%d') if book.publication_date else None,
            "stock_available": stock_available,
         })

    # Sort the book_details by review_count in descending order
    
    random.shuffle(book_details)
    book_details = book_details[:10]
    result_dict={
        "books":book_details,
        "genre":"Top Books"
    }
    return jsonify({'book_set': [result_dict], 'success': True})

@api_v2_books.route('/getGlobalBestsellersByAge', methods=['GET'])
def get_global_bestsellers_by_age():
    age = request.args.get('age')

    if not age or not age.isnumeric():
        return jsonify({'error': 'Invalid or missing age parameter'}), 400

    age = int(age)

    # Query the database to find books within the specified age group and with the "Global Bestseller" category.
    books = NewBook.query.filter(
        NewBook.min_age <= age,
        NewBook.max_age >= age,
        NewBook.categories.any(NewCategory.name == 'Global Bestseller')
    ).all()

    # Create a list of book details (name, author, rating, review count, ISBN, description).
    book_details = []
    for book in books:
        authors = book.authors.split(', ') if book.authors else [] 
        stock_available = 0
        book_record = Book.query.filter_by(isbn=book.isbn).first()
        if book_record:
            stock_available = book_record.stock_available
        else:
            stock_available=book.stock_available
        
        if stock_available:
            
         book_details.append({
            'name': book.name,
            'authors': authors,
            'rating': book.rating,
            'review_count': book.review_count,
            'isbn': book.isbn,
            'description': book.description,
            'image': book.image,
            'book_order': book.book_order,
            'publication_date': book.publication_date.strftime('%Y-%m-%d') if book.publication_date else None,
            "stock_available": stock_available,
         })
    random.shuffle(book_details)
    book_details = book_details[:10]
    result_dict={
        "books":book_details,
        "genre":"Top Books"
    }
    return jsonify({'book_set': [result_dict], 'success': True})

@api_v2_books.route('/getTeacherPicksByAge', methods=['GET'])
def get_teacher_picks_by_age():
    age = request.args.get('age')

    if not age or not age.isnumeric():
        return jsonify({'error': 'Invalid or missing age parameter'}), 400

    age = int(age)

    # Query the database to find books within the specified age group and with the "Teacher Pick" category.
    books = NewBook.query.filter(
        NewBook.min_age <= age,
        NewBook.max_age >= age,
        NewBook.categories.any(NewCategory.name == 'Teacher Pick')
    ).all()
    

    # Create a list of book details (name, author, rating, review count, ISBN, description).
    book_details = []
    for book in books:
        
        authors = book.authors.split(', ') if book.authors else []  
        stock_available = 0
        book_record = Book.query.filter_by(isbn=book.isbn).first()
        if book_record:
            stock_available = book_record.stock_available
        else:
            stock_available=book.stock_available
        
        if stock_available:
            
         book_details.append({
            'name': book.name,
            'authors': authors,
            'rating': book.rating,
            'review_count': book.review_count,
            'isbn': book.isbn,
            'description': book.description,
            'image': book.image,
            'book_order': book.book_order,
            'publication_date': book.publication_date.strftime('%Y-%m-%d') if book.publication_date else None,
            "stock_available": stock_available,
         })
    random.shuffle(book_details)
    book_details = book_details[:10]
    result_dict={
        "books":book_details,
        "genre":"Top Books"
    }
    return jsonify({'book_set': [result_dict], 'success': True})

@api_v2_books.route('/new-book-video', methods=['POST'])
def new_book_video():
    isbn = request.form.get('isbn')
    video_file = request.files.get('video')
    
    book = NewBook.query.filter_by(isbn=isbn).first()
    
    if not book:
        book =Book.query.filter_by(isbn=isbn).first()
        if not book:
          return jsonify({"success": False, "message": "Book with given ISBN does not exist"}), 404
    
    book_id = book.id
    
    extension = video_file.filename.split(".")[-1]
    upload_to_aws(video_file, 'book_videos/', f'book_videos/{isbn}.{extension}')
    s3_url = "https://bukrent-production.s3.ap-south-1.amazonaws.com"
    source = f'{s3_url}/book_videos/{isbn}.{extension}'
    
    NewBookVideo.create(source,book_id)  

    return jsonify({"success": True, "message": "Video uploaded and associated with the book"}), 200

@api_v2_books.route('/book-video/<isbn>', methods=['GET'])
def get_book_video(isbn):
    book = NewBook.query.filter_by(isbn=isbn).first()

    if not book:
        book = Book.query.filter_by(isbn=isbn).first()
        if not book:
            return jsonify({"success": False, "message": "Book with given ISBN does not exist"}), 404

    book_video = NewBookVideo.query.filter_by(book_id=book.id).first()

    if not book_video:
        return jsonify({"success": False, "message": "Video not found for the book"}), 404

    video_source = book_video.source
    return jsonify({"success": True, "video_source": video_source}), 200

@api_v2_books.route('/book-images/<isbn>', methods=['GET'])
def get_book_images(isbn):
    book = NewBook.query.filter_by(isbn=isbn).first()

    if not book:
        book = Book.query.filter_by(isbn=isbn).first()
        if not book:
            return jsonify({"success": False, "message": "Book with given ISBN does not exist"}), 404
    

    book_images_dict={
        "image":book.image
    }
    sources = NewBookImage.query.filter_by(book_id=book.id).all()
    source_info = []
    for source in sources:
        source_info.append({
            "image_id":source.id,
            "image_source": source.source,
            "image_angle":source.image_angle,
        })
    
    return jsonify({"success": True, "book_images": book_images_dict,"other_images":source_info}), 200  

@api_v2_books.route('/add-book-images/<isbn>', methods=['POST'])
def add_book_images(isbn):
    book = NewBook.query.filter_by(isbn=isbn).first()

    if not book:
        book = Book.query.filter_by(isbn=isbn).first()
        if not book:
            return jsonify({"success": False, "message": "Book with given ISBN does not exist"}), 404
    
    files = request.files.getlist('images')
    
    
    for file in files:
        try:
            extension = file.filename.split(".")[-1]
            upload_to_aws(file, 'secondary_book_images', f'secondary_book_images/{isbn}.{extension}')
            s3_url = "https://bukrent-production.s3.ap-south-1.amazonaws.com"
            image_source = f'{s3_url}/secondary_book_images/{isbn}.{extension}'
            NewBookImage.create(source=image_source, book_id=book.id)
        except Exception as e:
            return jsonify({"success": False, "message": str(e)}), 500
     
    return jsonify({"success": True, "message": "Images uploaded successfully"}), 200

@api_v2_books.route('/delete-book-image/<image_id>', methods=['GET'])
def delete_book_image(image_id):
    image = NewBookImage.query.get(image_id)

    if not image:
        return jsonify({"success": False, "message": "Image not found"}), 404

    # Delete the image
    image.delete()

    return jsonify({"success": True, "message": f"Image with ID {image_id} deleted successfully"}), 200

@api_v2_books.route('/change-book-angle/<image_id>', methods=['GET'])
def change_book_angle(image_id):
    new_angle = request.args.get('new_angle')  # Assuming the new angle is passed as a query parameter
    image = NewBookImage.query.get(image_id)

    if not image:
        return jsonify({"success": False, "message": "Image not found"}), 404

   
    image.image_angle = new_angle  # Update the image angle

    db.session.commit()  

    return jsonify({"success": True, "message": f"Image angle with ID {image_id} changed successfully"}), 200

@api_v2_books.route('/create_xlsx',methods=['GET'])
def create_xlsx():
  try:  
    iterator = 0
    books = NewBook.query.all()
    workbook = xlsxwriter.Workbook("./All_books.xlsx")
    worksheet = workbook.add_worksheet("All")
    worksheet.write(iterator, 0, "ISBN")
    worksheet.write(iterator, 1, "Name")
    worksheet.write(iterator, 2, "Review_Count")
    worksheet.write(iterator, 3, "Available")
    worksheet.write(iterator, 4, "Rentals")
    worksheet.write(iterator, 5, "Category")
    worksheet.write(iterator, 6, "Wishlist")
    worksheet.write(iterator, 7, "Previous")
    worksheet.write(iterator, 8, "Genres")
    worksheet.write(iterator, 9, "Rating")
    worksheet.write(iterator, 10, "Min_Age")
    worksheet.write(iterator, 11, "Max_Age")
    worksheet.write(iterator, 12, "Publisher")
    worksheet.write(iterator, 13, "Publication_Date")
    worksheet.write(iterator, 14, "Authors")
    worksheet.write(iterator, 15, "Video")
    worksheet.write(iterator, 16, "Primary Image")
    worksheet.write(iterator, 17, "Secondary Images")
    worksheet.write(iterator, 18, "Book Order")
    worksheet.write(iterator, 19, "Book Description")
    worksheet.write(iterator, 20, "Language")
    worksheet.write(iterator, 21, "hardcoverprice")
    worksheet.write(iterator, 22, "boardbookprice")
    worksheet.write(iterator, 23, "paperbackprice")
    

    for book in books:
        iterator += 1
        old_book = Book.query.filter_by(isbn=book.isbn).first()
        
        wishlist_count = 0
        previous_count = 0
        rentals=0
        available=0
        
        if old_book is not None:
           if not old_book.stock_available: 
             order = Order.query.filter(
                Order.book_id == old_book.id,
                cast(Order.placed_on, Date) >= cast(date.today() + timedelta(days=-7), Date)
             ).order_by(Order.placed_on).first()
            
             
        
           wishlist_count = Wishlist.query.filter_by(book_id=old_book.id).count()
           previous_count = Dump.query.filter_by(book_id=old_book.id, read_before=True).count() + \
                         Order.query.filter_by(book_id=old_book.id).count()
           rentals=old_book.rentals
           available=old_book.stock_available
        else:
            rentals=book.rentals
            available=book.stock_available   
           
        category_id = NewCategoryBook.query.filter_by(book_id=book.id).first()
        if category_id:
            category = NewCategory.query.filter_by(id=category_id.category_id).first()
            worksheet.write(iterator, 5, category.name)
        
        book_video = NewBookVideo.query.filter_by(book_id=book.id).first()
        
        if not book_video:
         worksheet.write(iterator, 15, "Video not found")
        else:
         
          worksheet.write(iterator, 15, book_video.source)
          
        sources = NewBookImage.query.filter_by(book_id=book.id).all()   
        source_values = [source.source for source in sources]
        result = ', '.join(source_values)
        
        worksheet.write(iterator, 17, result)        
        worksheet.write(iterator, 0, book.isbn)
        worksheet.write(iterator, 1, book.name)
        worksheet.write(iterator, 2, book.review_count)
        worksheet.write(iterator, 3, available)
        worksheet.write(iterator, 4, rentals)
        worksheet.write(iterator, 6, wishlist_count)
        worksheet.write(iterator, 7, previous_count)
        worksheet.write(iterator, 8, book.genre)
        worksheet.write(iterator, 9, book.rating)
        worksheet.write(iterator, 10, book.min_age)
        worksheet.write(iterator, 11, book.max_age)
        worksheet.write(iterator, 12, book.publisher)
        worksheet.write(iterator, 13, book.publication_date)
        worksheet.write(iterator, 14, book.authors)
        worksheet.write(iterator, 16, book.image)
        worksheet.write(iterator, 18, book.book_order)
        worksheet.write(iterator, 19, book.description)
        worksheet.write(iterator, 20, book.language)
        if book.hardcoverprice is not None: 
         worksheet.write(iterator, 21, book.hardcoverprice)
        if book.boardbookprice is not None: 
         worksheet.write(iterator, 22, book.boardbookprice)
        if book.paperbackprice is not None: 
         worksheet.write(iterator, 23, book.paperbackprice) 

        
        
    workbook.close()
    
    timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    file_path = "./All_books.xlsx"
    try:
     with open(file_path, 'rb') as file:
       
       upload_to_aws(file, 'excel_data/', f'excel_data/{timestamp}.{"xlsx"}')
       s3_url = "https://bukrent-production.s3.ap-south-1.amazonaws.com"
       file_url = f'{s3_url}/excel_data/{timestamp}.{"xlsx"}'
    
       return jsonify({"success": True, "message": "Excel file uploaded successfully", "url": file_url}), 200
        
    except Exception as e:
     return jsonify({"success": False, "message": str(e)}), 404    
  
  except Exception as e:
        return jsonify({"success": False, "message": e}), 404

@api_v2_books.route('/get-books-by-genre', methods=['GET'])
def get_books_by_age_and_genre():
    age = request.args.get('age')

    if not age:
        return jsonify({'error': 'Age not provided'}), 400

    try:
        age = int(age)
        if age < 0:
            raise ValueError("Age must be a non-negative integer")
    except ValueError:
        return jsonify({'error': 'Invalid age format'}), 400

    # Query the database to find books with the provided age and non-empty genre.
    books = db.session.query(NewBook).filter(
        NewBook.min_age <= age,
        NewBook.max_age >= age,
        NewBook.genre.isnot(None),
        NewBook.genre != "",
        NewBook.genre != "null"
    ).all()

    if not books:
        return jsonify({'error': 'No books found for the specified age and genre'}), 404

    # Organize books by genre
    books_by_genre = {}
    for book in books:
        if book.genre not in books_by_genre:
            books_by_genre[book.genre] = {'books': []}
            
        stock_available = 0
        book_record = Book.query.filter_by(isbn=book.isbn).first()
        if book_record:
            stock_available = book_record.stock_available
        else:
            stock_available=book.stock_available
        book_details = {
            'name': book.name,
            'isbn': book.isbn,
            'rating': book.rating,
            'max_age': book.max_age,
            'min_age': book.min_age,
            'review_count': book.review_count,
            'description': book.description,
            'image': book.image,
            "stock_available": stock_available,
        }
        books_by_genre[book.genre]['books'].append(book_details)

    # Convert to the desired response format
    response_data = {
        'book_set': [
            {'genre': genre, 'books': details['books']}
            for genre, details in books_by_genre.items()
        ]
    }

    return jsonify(response_data)